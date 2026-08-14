# Builds the MCAA rate SQLite DB that VANTAGE's takeoff consumes, from the SkySkraper
# rate-sheet Excel export. Two columns only: lookup_key (PRIMARY KEY) + manhours.
# Everything else in the workbook is key-building scaffold and is NOT exported.
# Columns are resolved by HEADER NAME, so this works on a 2-column export OR the full file.
#
# Usage:  python Scripts/export_rates_to_sqlite.py <input.xlsx> [output.db]
# Default output: Resources/cdx_weblem_rates.db (replaces the legacy Resources/RateSheet.json).

import openpyxl, sqlite3, sys, os

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_OUT = os.path.join(REPO_ROOT, "Resources", "cdx_weblem_rates.db")

def main():
    if len(sys.argv) < 2:
        sys.exit("Usage: python Scripts/export_rates_to_sqlite.py <input.xlsx> [output.db]")
    inp = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_OUT

    wb = openpyxl.load_workbook(inp, read_only=True, data_only=True)
    ws = wb["Rates"] if "Rates" in wb.sheetnames else wb.active
    rows = ws.iter_rows(values_only=True)

    header = [str(h).strip() if h is not None else "" for h in next(rows)]
    def col(name):
        if name not in header:
            sys.exit(f"Column '{name}' not found. Headers seen: {header}")
        return header.index(name)
    i_key, i_mh = col("lookup_key"), col("manhours")

    data, seen, blank_key, blank_mh = [], {}, 0, 0
    for r in rows:
        k = r[i_key]
        if k is None or str(k).strip() == "":
            blank_key += 1
            continue
        k = str(k)
        mh = r[i_mh]
        if mh is None or str(mh).strip() == "":
            blank_mh += 1
            mh_val = None
        else:
            try:
                mh_val = float(mh)
            except (ValueError, TypeError):
                mh_val = mh  # leave as-is; SQLite REAL affinity coerces numeric strings
        data.append((k, mh_val))
        seen[k] = seen.get(k, 0) + 1

    # Pre-flight: lookup_key must be unique or the PRIMARY KEY insert fails.
    dupes = {k: c for k, c in seen.items() if c > 1}
    if dupes:
        print(f"ABORT: {len(dupes)} duplicate lookup_key(s) found — PRIMARY KEY would fail.")
        print("Clear these in the workbook (DUPES column) before exporting:")
        for k, c in list(dupes.items())[:50]:
            print(f"  x{c}  {k}")
        if len(dupes) > 50:
            print(f"  ... and {len(dupes) - 50} more")
        sys.exit(1)

    if blank_key:
        print(f"NOTE: skipped {blank_key} row(s) with a blank lookup_key.")
    if blank_mh:
        print(f"WARNING: {blank_mh} row(s) have a blank manhours (stored as NULL).")

    if os.path.exists(out):
        os.remove(out)
    os.makedirs(os.path.dirname(out), exist_ok=True)
    con = sqlite3.connect(out)
    cur = con.cursor()
    cur.execute("CREATE TABLE rates (lookup_key TEXT PRIMARY KEY, manhours REAL)")
    cur.executemany("INSERT INTO rates (lookup_key, manhours) VALUES (?, ?)", data)
    con.commit()
    n = cur.execute("SELECT COUNT(*) FROM rates").fetchone()[0]
    con.close()
    print(f"Done: wrote {n} rows to {out}")

if __name__ == "__main__":
    main()
