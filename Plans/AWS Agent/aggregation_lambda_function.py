import json
import re
import boto3
import io
import logging
from collections import Counter
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger()
logger.setLevel(logging.INFO)

s3 = boto3.client("s3")

PROCESSING_BUCKET = "summit-takeoff-processing"


def lambda_handler(event, context):
    """
    Aggregation Lambda — combines per-drawing extraction JSONs into Excel output.
    Outputs 3 tabs: Material (all BOM items), Flagged (low/medium confidence items),
    Failed DWGs (drawings that failed extraction).
    All business logic (Labor generation, summary, rates) handled by app.
    """
    try:
        start_time = datetime.now(timezone.utc)
        
        if "extraction_keys" in event:
            extraction_keys = event["extraction_keys"]
            batch_id = event.get("batch_id", f"test-{start_time.strftime('%Y%m%d-%H%M%S')}")
            logger.info(f"Test mode: {len(extraction_keys)} extraction keys provided")
        elif "batch_id" in event:
            batch_id = event["batch_id"]
            extraction_keys = load_batch_extraction_keys(batch_id)
            logger.info(f"Batch mode: {len(extraction_keys)} extractions found for {batch_id}")
        else:
            raise ValueError("Event must contain either 'extraction_keys' or 'batch_id'")
        
        extractions = []
        for key in extraction_keys:
            extraction = load_extraction(key)
            if extraction:
                extractions.append(extraction)
        
        logger.info(f"Loaded {len(extractions)} successful extractions")
        
        if not extractions:
            raise ValueError("No valid extractions to aggregate")
        
        # Consensus backfill: use batch-wide majority to fill missing class_rating values
        consensus_backfill_class_rating(extractions)
        
        material_rows = build_material_rows(extractions)
        flagged_rows = build_flagged_rows(extractions)
        failed_rows = load_failure_markers(batch_id)
        
        logger.info(f"Built {len(material_rows)} material rows, {len(flagged_rows)} flagged rows, "
                    f"{len(failed_rows)} failed drawings")
        
        excel_bytes = generate_excel(material_rows, flagged_rows, failed_rows)
        
        excel_key = f"batches/{batch_id}/output/takeoff_{batch_id}.xlsx"
        s3.put_object(
            Bucket=PROCESSING_BUCKET,
            Key=excel_key,
            Body=excel_bytes,
            ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        logger.info(f"Excel written to s3://{PROCESSING_BUCKET}/{excel_key}")
        
        elapsed_ms = int((datetime.now(timezone.utc) - start_time).total_seconds() * 1000)
        
        return {
            "status": "completed",
            "batch_id": batch_id,
            "excel_path": excel_key,
            "total_drawings": len(extractions),
            "total_material_rows": len(material_rows),
            "total_flagged_rows": len(flagged_rows),
            "total_failed_drawings": len(failed_rows),
            "processing_time_ms": elapsed_ms
        }
        
    except Exception as e:
        logger.error(f"Aggregation failed: {str(e)}", exc_info=True)
        return {
            "status": "failed",
            "error": str(e)
        }


# ---------------------------------------------------------------------------
# Consensus backfill
# ---------------------------------------------------------------------------

def consensus_backfill_class_rating(extractions):
    """
    Batch-wide consensus backfill for class_rating.
    Groups all BOM items by raw description. For each description group,
    if any items have class_rating and others don't, the majority value
    is used to fill the gaps. Mutates bom_items in place.
    """
    # Pass 1: collect all class_rating values grouped by raw description
    desc_ratings = {}
    for extraction in extractions:
        for item in extraction.get("bom_items", []):
            desc = item.get("description", "")
            if not desc:
                continue
            if desc not in desc_ratings:
                desc_ratings[desc] = []
            rating = item.get("class_rating")
            if rating:
                desc_ratings[desc].append(str(rating))

    # Build majority lookup — only for descriptions that have at least one rating
    majority_lookup = {}
    for desc, ratings in desc_ratings.items():
        if not ratings:
            continue
        counter = Counter(ratings)
        winner, count = counter.most_common(1)[0]
        majority_lookup[desc] = winner
        # Log conflicts where the model returned different values for the same description
        if len(counter) > 1:
            logger.warning(f"CLASS_RATING_CONFLICT: Description '{desc[:80]}...' has conflicting values: "
                           f"{dict(counter)} — using majority '{winner}'")

    # Pass 2: backfill missing class_rating values
    backfill_count = 0
    for extraction in extractions:
        for item in extraction.get("bom_items", []):
            if item.get("class_rating"):
                continue
            desc = item.get("description", "")
            if desc in majority_lookup:
                item["class_rating"] = majority_lookup[desc]
                backfill_count += 1

    if backfill_count > 0:
        logger.info(f"CLASS_RATING_BACKFILL: Filled {backfill_count} missing values "
                    f"across {len(majority_lookup)} unique descriptions using batch consensus")


# ---------------------------------------------------------------------------
# Extraction loading
# ---------------------------------------------------------------------------

def load_batch_extraction_keys(batch_id):
    prefix = f"batches/{batch_id}/extractions/"
    paginator = s3.get_paginator("list_objects_v2")
    keys = []
    
    for page in paginator.paginate(Bucket=PROCESSING_BUCKET, Prefix=prefix):
        for obj in page.get("Contents", []):
            key = obj["Key"]
            if key.endswith(".json"):
                keys.append(key)
    
    return keys


def load_extraction(key):
    try:
        response = s3.get_object(Bucket=PROCESSING_BUCKET, Key=key)
        data = json.loads(response["Body"].read().decode("utf-8"))
        
        if "extraction" in data and "status" in data:
            if data["status"] != "success":
                logger.warning(f"Skipping failed extraction: {key}")
                return None
            extraction = data["extraction"]
            extraction["_source_key"] = data.get("source_key", key)
        else:
            extraction = data
            extraction["_source_key"] = key
        
        return extraction
        
    except Exception as e:
        logger.error(f"Failed to load extraction {key}: {e}")
        return None


# ---------------------------------------------------------------------------
# Failure marker loading
# ---------------------------------------------------------------------------

def load_failure_markers(batch_id):
    """
    Load failure markers written by the extraction Lambda for drawings that failed.
    Returns list of dicts suitable for writing to the Failed DWGs tab.
    Safe to return [] if prefix is empty or doesn't exist.
    """
    prefix = f"batches/{batch_id}/failures/"
    paginator = s3.get_paginator("list_objects_v2")
    failure_rows = []
    
    try:
        for page in paginator.paginate(Bucket=PROCESSING_BUCKET, Prefix=prefix):
            for obj in page.get("Contents", []):
                key = obj["Key"]
                if not key.endswith(".json"):
                    continue
                try:
                    response = s3.get_object(Bucket=PROCESSING_BUCKET, Key=key)
                    data = json.loads(response["Body"].read().decode("utf-8"))
                    failure_rows.append({
                        "drawing_name": data.get("drawing_name", ""),
                        "source_key": data.get("source_key", ""),
                        "error": data.get("error", ""),
                        "timestamp": data.get("timestamp", ""),
                    })
                except Exception as e:
                    logger.error(f"Failed to load failure marker {key}: {e}")
    except Exception as e:
        logger.warning(f"Could not list failure markers at {prefix}: {e}")
    
    return failure_rows


# ---------------------------------------------------------------------------
# Data cleaning
# ---------------------------------------------------------------------------

def clean_quantity(qty):
    """
    Convert quantity to decimal number, stripping ' and " marks.
    "41.3'" -> 41.3
    "12" -> 12
    "5.5\"" -> 5.5
    """
    if qty is None:
        return None
    
    qty_str = str(qty).replace("'", "").replace('"', "").strip()
    
    try:
        return float(qty_str)
    except (ValueError, TypeError):
        return qty  # Return original if can't parse


def normalize_size(raw_size):
    """Convert raw size string from drawing to decimal format.
    Handles fractions, mixed numbers, and reducing sizes with X separator.
    Examples: "3/4" -> "0.75", "1-1/2" -> "1.5", "3/4X1/2" -> "0.75x0.5"
    """
    if not raw_size:
        return raw_size

    s = str(raw_size).strip()

    # Handle reducing sizes — split on X/x with optional spaces
    if re.search(r'[xX]', s):
        parts = re.split(r'\s*[xX]\s*', s, maxsplit=1)
        if len(parts) == 2:
            return "x".join(normalize_single_size(p.strip()) for p in parts)

    return normalize_single_size(s)


def normalize_single_size(s):
    """Convert a single size value (whole, fraction, or mixed) to decimal string."""
    s = s.strip()

    # Already a plain number — pass through
    try:
        float(s)
        return s
    except ValueError:
        pass

    # Mixed number: 1-1/2, 1 1/2, 2-3/4, 2 3/4, 10-3/4, etc.
    mixed_match = re.match(r'^(\d+)[-\s](\d+)/(\d+)$', s)
    if mixed_match:
        whole = int(mixed_match.group(1))
        num = int(mixed_match.group(2))
        den = int(mixed_match.group(3))
        if den != 0:
            result = whole + num / den
            return format_decimal(result)

    # Pure fraction: 3/4, 1/2, etc.
    frac_match = re.match(r'^(\d+)/(\d+)$', s)
    if frac_match:
        num = int(frac_match.group(1))
        den = int(frac_match.group(2))
        if den != 0:
            return format_decimal(num / den)

    return s


def format_decimal(value):
    """Format a float to clean decimal string — no trailing zeros."""
    formatted = f"{value:.4f}".rstrip('0').rstrip('.')
    return formatted


# ---------------------------------------------------------------------------
# Row builders
# ---------------------------------------------------------------------------

def build_material_rows(extractions):
    """
    Build material rows — one row per BOM item.
    ShopField hardcoded to 1. C# app assigns real Shop/Field values post-download.
    """
    rows = []

    for extraction in extractions:
        drawing_number = extraction.get("drawing_number", "UNKNOWN")
        title_block = {k.rstrip(":").strip(): v for k, v in extraction.get("title_block", {}).items()}
        bom_items = extraction.get("bom_items", [])
        
        for item in bom_items:
            conn_type = item.get("connection_type") or ""
            conn_qty = item.get("connection_qty", 0) or 0
            component = item.get("component") or ""

            shop_field = 1  # C# post-processing assigns real values
            
            # Build concatenated description
            # Format: size IN - component - thickness - class - pipe spec - material - length
            size = item.get("size") or ""
            thickness = item.get("thickness") or ""
            matl_grp = item.get("matl_grp") or ""
            component = item.get("component") or ""
            commodity_code = item.get("commodity_code") or ""
            class_rating = item.get("class_rating") or ""
            length = item.get("length") or ""

            normalized_conn_size = normalize_size(item.get("connection_size"))
            normalized_size = normalize_size(size)

            row = {
                "drawing_number": drawing_number,
                "item_id": item.get("item_id"),
                "component": component,
                "size": normalized_size,
                "raw_description": item.get("description"),
                "quantity": clean_quantity(item.get("quantity")),
                "connection_qty": conn_qty,
                "connection_type": conn_type,
                "connection_size": normalized_conn_size,
                "thickness": thickness,
                "class_rating": item.get("class_rating"),
                "length": item.get("length"),
                "matl_grp": item.get("matl_grp"),
                "matl_grp_desc": item.get("matl_grp_desc"),
                "commodity_code": commodity_code,
                "shop_field": shop_field,
                "confidence": item.get("confidence"),
                "flag": item.get("flag"),
                **{f"tb_{k}": v for k, v in title_block.items()}
            }
            rows.append(row)
    
    return rows


def build_flagged_rows(extractions):
    rows = []
    
    for extraction in extractions:
        drawing_number = extraction.get("drawing_number", "UNKNOWN")
        title_block = {k.rstrip(":").strip(): v for k, v in extraction.get("title_block", {}).items()}
        bom_items = extraction.get("bom_items", [])
        
        for item in bom_items:
            confidence = item.get("confidence", "high")
            
            if confidence.lower() not in ("low", "medium"):
                continue
            
            row = {
                "drawing_number": drawing_number,
                "item_id": item.get("item_id"),
                "size": normalize_size(item.get("size")),
                "description": item.get("description"),
                "component": item.get("component"),
                "connection_qty": item.get("connection_qty"),
                "connection_type": item.get("connection_type"),
                "matl_grp": item.get("matl_grp"),
                "matl_grp_desc": item.get("matl_grp_desc"),
                "thickness": item.get("thickness"),
                "class_rating": item.get("class_rating"),
                "confidence": confidence,
                "flag": item.get("flag"),
                "override_component": "",
                "override_connection_qty": "",
                "override_connection_type": "",
                "override_notes": ""
            }
            rows.append(row)
    
    return rows


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------

def generate_excel(material_rows, flagged_rows, failed_rows):
    wb = Workbook()
    
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    ws_material = wb.active
    ws_material.title = "Material"
    write_material_tab(ws_material, material_rows, header_font, header_fill, thin_border)
    
    ws_flagged = wb.create_sheet("Flagged")
    write_flagged_tab(ws_flagged, flagged_rows, header_font, header_fill, thin_border)
    
    ws_failed = wb.create_sheet("Failed DWGs")
    write_failed_tab(ws_failed, failed_rows, header_font, header_fill, thin_border)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def write_material_tab(ws, material_rows, header_font, header_fill, thin_border):
    if not material_rows:
        ws.cell(row=1, column=1, value="No material data")
        return
    
    tb_fields = set()
    for row in material_rows:
        for key in row.keys():
            if key.startswith("tb_"):
                tb_fields.add(key)
    tb_fields = sorted(list(tb_fields))
    
    fixed_columns = [
        ("drawing_number", "Drawing Number"),
        ("item_id", "Item ID"),
        ("component", "Component"),
        ("size", "Size"),
        ("raw_description", "Raw Description"),
        ("quantity", "Quantity"),
        ("connection_qty", "Connection Qty"),
        ("connection_type", "Connection Type"),
        ("connection_size", "Connection Size"),
        ("thickness", "Thickness"),
        ("class_rating", "Class Rating"),
        ("length", "Length"),
        ("matl_grp", "Matl_Grp"),
        ("matl_grp_desc", "Matl_Grp_Desc"),
        ("commodity_code", "Commodity Code"),
        ("shop_field", "ShopField"),
        ("confidence", "Confidence"),
        ("flag", "Flag"),
    ]
    
    columns = fixed_columns + [(tb, tb.replace("tb_", "")) for tb in tb_fields]
    
    for col_idx, (key, label) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    for row_idx, row_data in enumerate(material_rows, start=2):
        for col_idx, (key, label) in enumerate(columns, start=1):
            value = row_data.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
    
    for col_idx, (key, label) in enumerate(columns, start=1):
        max_len = len(label)
        for row_data in material_rows[:100]:
            val_len = len(str(row_data.get(key, "")))
            if val_len > max_len:
                max_len = val_len
        width = min(max(max_len + 2, 10), 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    ws.freeze_panes = "A2"


def write_flagged_tab(ws, flagged_rows, header_font, header_fill, thin_border):
    if not flagged_rows:
        ws.cell(row=1, column=1, value="No flagged items — all extractions high confidence")
        return
    
    columns = [
        ("drawing_number", "Drawing Number"),
        ("item_id", "Item ID"),
        ("size", "Size"),
        ("description", "Description"),
        ("component", "Component"),
        ("connection_qty", "Connection Qty"),
        ("connection_type", "Connection Type"),
        ("matl_grp", "Matl_Grp"),
        ("matl_grp_desc", "Matl_Grp_Desc"),
        ("thickness", "Thickness"),
        ("class_rating", "Class Rating"),
        ("confidence", "Confidence"),
        ("flag", "Flag Reason"),
        ("override_component", "Override Component"),
        ("override_connection_qty", "Override Conn Qty"),
        ("override_connection_type", "Override Conn Type"),
        ("override_notes", "Override Notes"),
    ]
    
    override_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    
    for col_idx, (key, label) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill if not key.startswith("override_") else override_fill
        cell.border = thin_border
    
    for row_idx, row_data in enumerate(flagged_rows, start=2):
        for col_idx, (key, label) in enumerate(columns, start=1):
            value = row_data.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if key.startswith("override_"):
                cell.fill = override_fill
    
    for col_idx, (key, label) in enumerate(columns, start=1):
        max_len = len(label)
        for row_data in flagged_rows[:100]:
            val_len = len(str(row_data.get(key, "")))
            if val_len > max_len:
                max_len = val_len
        width = min(max(max_len + 2, 10), 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    ws.freeze_panes = "A2"


def write_failed_tab(ws, failed_rows, header_font, header_fill, thin_border):
    if not failed_rows:
        ws.cell(row=1, column=1, value="No failed drawings — all drawings extracted successfully")
        return
    
    columns = [
        ("drawing_name", "Drawing Name"),
        ("source_key", "Source Key"),
        ("error", "Error"),
        ("timestamp", "Timestamp (UTC)"),
    ]
    
    for col_idx, (key, label) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    for row_idx, row_data in enumerate(failed_rows, start=2):
        for col_idx, (key, label) in enumerate(columns, start=1):
            value = row_data.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
    
    for col_idx, (key, label) in enumerate(columns, start=1):
        max_len = len(label)
        for row_data in failed_rows[:100]:
            val_len = len(str(row_data.get(key, "")))
            if val_len > max_len:
                max_len = val_len
        width = min(max(max_len + 2, 10), 80)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    ws.freeze_panes = "A2"
